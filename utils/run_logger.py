from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


EXCEL_HEADERS = [
    "run_id",
    "timestamp",
    "experiment_name",
    "seed",
    "size",
    "full_obs",
    "n_envs",
    "total_timesteps",
    "learning_rate",
    "n_steps",
    "batch_size",
    "n_epochs",
    "gamma",
    "gae_lambda",
    "clip_range",
    "ent_coef",
    "vf_coef",
    "max_grad_norm",
    "key_bonus",
    "door_bonus",
    "features_dim",
    "pi_layers",
    "vf_layers",
    "param_count",
    "status",
    "best_mean_reward",
    "last_success_rate",
    "final_model_path",
    "best_model_path",
    "notes",
]


def ensure_workbook(path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "runs"
    ws.append(EXCEL_HEADERS)
    wb.save(path)


def append_run(path: str | Path, row_data: dict[str, Any]) -> int:
    path = Path(path)
    ensure_workbook(path)

    wb = load_workbook(path)
    ws = wb["runs"]

    row = [row_data.get(col, "") for col in EXCEL_HEADERS]
    ws.append(row)

    row_idx = ws.max_row
    wb.save(path)
    return row_idx


def update_run(path: str | Path, row_idx: int, updates: dict[str, Any]) -> None:
    path = Path(path)
    wb = load_workbook(path)
    ws = wb["runs"]

    header_to_col = {
        ws.cell(row=1, column=col_idx).value: col_idx
        for col_idx in range(1, ws.max_column + 1)
    }

    for key, value in updates.items():
        if key in header_to_col:
            ws.cell(row=row_idx, column=header_to_col[key], value=value)

    wb.save(path)